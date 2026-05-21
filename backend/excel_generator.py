import io
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def clean_json(s: str) -> str:
    s = s.strip()
    s = re.sub(r"^```(?:json)?\s*", "", s)
    s = re.sub(r"\s*```$", "", s)
    return s.strip()


def _header_row(ws, row: int, headers: list[str], bg: str = "1A3A5C"):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color="FFFFFF", size=11)
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26


def _title_row(ws, text: str, span: int, row: int = 1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0D2137")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32


def _parse_gherkin(text: str) -> dict:
    given, when, then, current = [], [], [], None
    for line in text.split("\n"):
        line = line.strip()
        if not line or line.lower().startswith("scenario"):
            continue
        low = line.lower()
        if low.startswith("given "):
            current = "given"; given.append(line)
        elif low.startswith("when "):
            current = "when"; when.append(line)
        elif low.startswith("then "):
            current = "then"; then.append(line)
        elif low.startswith("and ") or low.startswith("but "):
            if current == "given": given.append(line)
            elif current == "when": when.append(line)
            elif current == "then": then.append(line)
    return {
        "given": "\n".join(given),
        "when": "\n".join(when),
        "then": "\n".join(then),
    }


TYPE_COLORS = {
    "Positive":            "D4EDDA",
    "Negative":            "F8D7DA",
    "Validation":          "FFF3CD",
    "Authorization":       "D1ECF1",
    "Dependency":          "E2D9F3",
    "Workflow Transition": "FCE8D5",
    "Edge Case":           "D6EAF8",
    "Cross Module":        "FDEBD0",
    "Error Handling":      "FADBD8",
    "Audit Validation":    "D5F5E3",
}

STATUS_COLORS = {
    "Good":               "D4EDDA",
    "Pass":               "D4EDDA",
    "Partial":            "FFF3CD",
    "Needs Improvement":  "F8D7DA",
    "Fail":               "F8D7DA",
}


def _sheet_scenarios(wb: Workbook, final_output: str):
    ws = wb.active
    ws.title = "Test Scenarios"

    try:
        data = json.loads(clean_json(final_output))
    except Exception:
        ws.append(["Could not parse scenario data."])
        return

    use_case = data.get("use_case", "")
    scenarios = data.get("gherkin_scenarios", [])

    _title_row(ws, f"Test Scenarios — {use_case}", 7)
    _header_row(ws, 2, ["#", "Type", "Scenario Name", "Given (Preconditions)", "When (Action)", "Then (Expected Result)", "Full Gherkin"])

    for i, s in enumerate(scenarios, 1):
        steps = _parse_gherkin(s.get("gherkin", ""))
        stype = s.get("scenario_type", "")
        color = TYPE_COLORS.get(stype, "FFFFFF")
        fill = PatternFill("solid", fgColor=color)
        row_num = i + 2

        values = [i, stype, s.get("scenario_name", ""), steps["given"], steps["when"], steps["then"], s.get("gherkin", "")]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       horizontal="center" if col == 1 else "left")
        ws.row_dimensions[row_num].height = 70

    for col, width in zip(range(1, 8), [5, 22, 36, 42, 32, 42, 62]):
        ws.column_dimensions[get_column_letter(col)].width = width


def _sheet_review(wb: Workbook, review_feedback: str):
    ws = wb.create_sheet("Review Summary")

    try:
        data = json.loads(clean_json(review_feedback))
    except Exception:
        ws.append(["Could not parse review data."])
        return

    _title_row(ws, "Review Summary", 2)
    _header_row(ws, 2, ["Coverage Area", "Status"])

    row = 3
    for area, status in data.get("coverage_summary", {}).items():
        label = area.replace("_", " ").title()
        color = STATUS_COLORS.get(status, "FFFFFF")
        ws.cell(row=row, column=1, value=label).alignment = Alignment(vertical="center")
        cell = ws.cell(row=row, column=2, value=status)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    overall = data.get("overall_review_status", "")
    ws.cell(row=row, column=1, value="Overall Status").font = Font(bold=True, size=12)
    cell = ws.cell(row=row, column=2, value=overall)
    cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(overall, "EEEEEE"))
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26

    sections = [
        ("Missing Scenarios",       data.get("missing_scenarios", [])),
        ("Weak Scenarios",          data.get("weak_scenarios", [])),
        ("Duplicate Scenarios",     data.get("duplicate_scenarios", [])),
        ("Gherkin Issues",          data.get("gherkin_issues", [])),
        ("Recommended Improvements",data.get("recommended_improvements", [])),
    ]

    for title, items in sections:
        if not items:
            continue
        row += 2
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, color="0D2137")
        row += 1
        for item in items:
            ws.cell(row=row, column=1, value=f"• {item}")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 22
            row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22


def _sheet_categories(wb: Workbook, final_output: str):
    """Build the Scenario Categories sheet from the Gherkin output so it always
    matches the Test Scenarios sheet (both derived from the same source)."""
    ws = wb.create_sheet("Scenario Categories")

    try:
        data = json.loads(clean_json(final_output))
    except Exception:
        ws.append(["Could not parse scenario category data."])
        return

    use_case = data.get("use_case", "")
    scenarios = data.get("gherkin_scenarios", [])

    _title_row(ws, f"Scenario Categories — {use_case}", 2)

    TYPE_ORDER = [
        ("Positive",            "D4EDDA"),
        ("Negative",            "F8D7DA"),
        ("Validation",          "FFF3CD"),
        ("Authorization",       "D1ECF1"),
        ("Dependency",          "E2D9F3"),
        ("Workflow Transition",  "FCE8D5"),
        ("Edge Case",           "D6EAF8"),
        ("Cross Module",        "FDEBD0"),
        ("Error Handling",      "FADBD8"),
        ("Audit Validation",    "D5F5E3"),
    ]

    # Group scenario names by type preserving the canonical order
    grouped: dict[str, list[str]] = {}
    for s in scenarios:
        stype = s.get("scenario_type", "Other")
        grouped.setdefault(stype, []).append(s.get("scenario_name", ""))

    row = 2
    for label, color in TYPE_ORDER:
        items = grouped.get(label, [])
        if not items:
            continue
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=f"{label} Scenarios")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2B5F8E")
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1

        fill = PatternFill("solid", fgColor=color)
        for idx, name in enumerate(items, 1):
            ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
            cell = ws.cell(row=row, column=2, value=name)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 32
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 90


def generate_excel(question: str, final_output: str, review_feedback: str, generated_scenarios: str) -> bytes:
    wb = Workbook()
    _sheet_scenarios(wb, final_output)
    _sheet_review(wb, review_feedback)
    _sheet_categories(wb, final_output)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
